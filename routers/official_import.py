from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from pydantic import BaseModel
from typing import Optional
from io import BytesIO
from datetime import date
import csv

from openpyxl import load_workbook
from database import get_db

from .import_match import (
    score_person_candidate, band, diff_fields, _norm as _m_norm
)

router = APIRouter(prefix="/officials/import", tags=["official-import"])

MAX_PREVIEW_ROWS = 1000
MAX_FILE_BYTES = 5_000_000

FIELD_CHOICES = [
    "jurisdiction_name", "name", "title", "role_type",
    "email", "phone", "fax",
    "mailing_address", "physical_address",
    "source",
]

GUESS_PATTERNS = {
    "jurisdiction_name": ["jurisdiction", "entity", "city", "county", "district", "agency"],
    "name": ["name", "official", "person", "full name"],
    "title": ["title", "position", "role", "office"],
    "email": ["email", "e-mail", "mail"],
    "phone": ["phone", "telephone", "tel"],
    "fax": ["fax"],
    "mailing_address": ["mailing", "mail address", "po box"],
    "physical_address": ["physical", "street", "address"],
    "source": ["source"],
    "role_type": ["role type", "elected", "staff"],
}

DIFF_FIELDS = ["name", "title", "email", "phone", "fax",
               "mailing_address", "physical_address", "role_type"]


def _parse_upload(filename: str, content: bytes):
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        first = next(it, None) or []
        headers = [str(h).strip() if h is not None else "" for h in first]
        data = []
        for row in it:
            padded = list(row) + [None] * (len(headers) - len(row))
            data.append([("" if v is None else str(v)).strip() for v in padded[:len(headers)]])
            if len(data) >= MAX_PREVIEW_ROWS:
                break
        return headers, data
    if name.endswith(".csv") or name.endswith(".txt"):
        txt = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(txt.splitlines())
        headers = [h.strip() for h in next(reader, [])]
        data = []
        for row in reader:
            padded = list(row) + [""] * (len(headers) - len(row))
            data.append([c.strip() for c in padded[:len(headers)]])
            if len(data) >= MAX_PREVIEW_ROWS:
                break
        return headers, data
    from .import_vision import extract_via_vision, _detect_media_type
    if _detect_media_type(filename):
        return extract_via_vision(filename, content, "official")
    raise HTTPException(400, "Unsupported file type. Use .csv, .xlsx, .pdf, .jpg, or .png")


def _guess_mapping(headers):
    mapping = {}
    used = set()
    for field, pats in GUESS_PATTERNS.items():
        best = None
        for h in headers:
            if h in used or not h:
                continue
            hl = h.lower().strip()
            if any(p == hl for p in pats):
                best = h
                break
        if not best:
            for h in headers:
                if h in used or not h:
                    continue
                hl = h.lower().strip()
                if any(p in hl for p in pats):
                    best = h
                    break
        if best:
            mapping[field] = best
            used.add(best)
    return mapping


@router.post("/preview")
async def preview(file: UploadFile = File(...)):
    content = await file.read()
    if len(content) > MAX_FILE_BYTES:
        raise HTTPException(413, "File too large (max 5MB)")
    headers, rows = _parse_upload(file.filename, content)
    return {
        "columns": headers,
        "rows": rows,
        "suggested_mapping": _guess_mapping(headers),
        "field_choices": FIELD_CHOICES,
        "truncated": len(rows) >= MAX_PREVIEW_ROWS,
    }


class DiffRequest(BaseModel):
    columns: list[str]
    rows: list[list[str]]
    mapping: dict[str, str]


async def _resolve_jurisdiction(db: AsyncSession, jname: str):
    """Return (exact_jid, exact_candidates, fuzzy_candidates)."""
    if not jname:
        return None, [], []
    n = _m_norm(jname)
    r = await db.execute(text(
        "SELECT jurisdiction_id, name, type FROM common.jurisdictions "
        "WHERE lower(trim(name)) = :n LIMIT 1"
    ), {"n": n})
    row = r.mappings().first()
    exact = dict(row) if row else None

    # Also grab fuzzy candidates — "Madison County Sheriff's Office" should find "Madison County"
    pattern = "%" + n.replace("%", "") + "%"
    r = await db.execute(text(
        "SELECT jurisdiction_id, name, type FROM common.jurisdictions "
        "WHERE lower(name) ILIKE :p OR lower(:n) ILIKE '%' || lower(name) || '%' "
        "ORDER BY length(name) LIMIT 10"
    ), {"p": pattern, "n": n})
    fuzzy = [dict(x) for x in r.mappings().all()]

    # Strip exact from fuzzy list
    if exact:
        fuzzy = [c for c in fuzzy if c["jurisdiction_id"] != exact["jurisdiction_id"]]

    return exact, ([exact] if exact else []), fuzzy


async def _fetch_officials_for_jids(db: AsyncSession, jids: list[int]):
    if not jids:
        return []
    r = await db.execute(text(
        "SELECT official_id, jurisdiction_id, name, title, email, phone, fax, "
        "       mailing_address, physical_address, role_type, source "
        "FROM public.officials "
        "WHERE jurisdiction_id = ANY(:jids) AND ended_date IS NULL"
    ), {"jids": jids})
    return [dict(x) for x in r.mappings().all()]


@router.post("/diff")
async def diff(payload: DiffRequest, db: AsyncSession = Depends(get_db)):
    cols = payload.columns
    mapping = payload.mapping or {}
    col_index = {h: i for i, h in enumerate(cols)}

    required = ["jurisdiction_name", "name", "title"]
    for f in required:
        if f not in mapping or mapping[f] not in col_index:
            raise HTTPException(400, f"Mapping required: {f}")

    def cell(row, field):
        col = mapping.get(field)
        if not col or col not in col_index:
            return None
        i = col_index[col]
        if i >= len(row):
            return None
        v = (row[i] or "").strip()
        return v or None

    juris_cache = {}
    results = []

    for idx, row in enumerate(payload.rows):
        jname = cell(row, "jurisdiction_name")
        name = cell(row, "name")
        title = cell(row, "title")
        incoming = {f: cell(row, f) for f in FIELD_CHOICES}

        if not name or not title:
            results.append({
                "row_index": idx,
                "status": "ERROR",
                "error": "Missing name or title",
                "incoming": incoming,
            })
            continue

        ckey = _m_norm(jname)
        if ckey in juris_cache:
            exact, exact_list, fuzzy_list = juris_cache[ckey]
        else:
            exact, exact_list, fuzzy_list = await _resolve_jurisdiction(db, jname)
            juris_cache[ckey] = (exact, exact_list, fuzzy_list)

        # Gather candidate jurisdiction IDs
        exact_ids = [c["jurisdiction_id"] for c in exact_list]
        fuzzy_ids = [c["jurisdiction_id"] for c in fuzzy_list]
        all_jids = exact_ids + fuzzy_ids

        if not all_jids:
            results.append({
                "row_index": idx,
                "status": "UNMATCHED",
                "incoming": incoming,
                "jurisdiction_candidates": [],
            })
            continue

        officials = await _fetch_officials_for_jids(db, all_jids)

        # Score every candidate
        exact_id_set = set(exact_ids)
        scored = []
        for o in officials:
            jm = "exact" if o["jurisdiction_id"] in exact_id_set else "fuzzy"
            score, parts = score_person_candidate(incoming, o, jm)
            scored.append({"score": score, "parts": parts, "official": o})
        scored.sort(key=lambda s: s["score"], reverse=True)

        top = scored[0] if scored else None
        top_score = top["score"] if top else 0
        status_band = band(top_score, person=True) if top else "NEW"

        top_candidates = [
            {
                "official_id": s["official"]["official_id"],
                "name": s["official"]["name"],
                "title": s["official"]["title"],
                "jurisdiction_id": s["official"]["jurisdiction_id"],
                "score": s["score"],
                "score_parts": s["parts"],
                "existing": s["official"],
            }
            for s in scored[:5]
        ]

        # Resolve target jurisdiction_id for insert (exact > best fuzzy > None)
        target_jid = exact["jurisdiction_id"] if exact else (fuzzy_list[0]["jurisdiction_id"] if fuzzy_list else None)

        result = {
            "row_index": idx,
            "incoming": incoming,
            "jurisdiction_id": target_jid,
            "jurisdiction_exact": exact,
            "jurisdiction_candidates": fuzzy_list[:5],
            "candidates": top_candidates,
        }

        if status_band == "AUTO" and top:
            ex = top["official"]
            fill, overwrite = diff_fields(incoming, ex, DIFF_FIELDS)
            all_changes = fill + overwrite
            if not all_changes:
                result["status"] = "SAME"
                result["best_match"] = top_candidates[0]
                result["existing"] = ex
            else:
                result["status"] = "CHANGED"
                result["best_match"] = top_candidates[0]
                result["existing"] = ex
                result["fill_fields"] = fill
                result["overwrite_fields"] = overwrite
        elif status_band == "POSSIBLE" and top:
            result["status"] = "POSSIBLE"
            result["best_match"] = top_candidates[0]
        else:
            result["status"] = "NEW"

        results.append(result)

    summary = {}
    for r in results:
        summary[r["status"]] = summary.get(r["status"], 0) + 1

    return {"rows": results, "summary": summary}


class CommitDecision(BaseModel):
    row_index: int
    action: str  # "insert" | "merge" | "skip"
    jurisdiction_id: Optional[int] = None
    existing_official_id: Optional[int] = None
    approved_fields: Optional[list[str]] = None  # for action="merge"


class CommitRequest(BaseModel):
    columns: list[str]
    rows: list[list[str]]
    mapping: dict[str, str]
    decisions: list[CommitDecision]
    source: Optional[str] = None


@router.post("/commit")
async def commit(payload: CommitRequest, db: AsyncSession = Depends(get_db)):
    cols = payload.columns
    mapping = payload.mapping or {}
    col_index = {h: i for i, h in enumerate(cols)}

    def cell(row, field):
        c = mapping.get(field)
        if not c or c not in col_index:
            return None
        i = col_index[c]
        if i >= len(row):
            return None
        v = (row[i] or "").strip()
        return v or None

    inserted = 0
    merged = 0
    skipped = 0
    errors = []
    today = date.today()

    for d in payload.decisions:
        if d.action == "skip":
            skipped += 1
            continue
        if d.row_index < 0 or d.row_index >= len(payload.rows):
            errors.append({"row_index": d.row_index, "error": "row out of range"})
            continue
        row = payload.rows[d.row_index]
        name = cell(row, "name")
        title = cell(row, "title")

        if d.action == "merge":
            if not d.existing_official_id:
                errors.append({"row_index": d.row_index, "error": "merge needs existing_official_id"})
                continue
            approved = set(d.approved_fields or [])
            if not approved:
                skipped += 1
                continue
            allowed = set(DIFF_FIELDS)
            sets = []
            params = {"oid": d.existing_official_id}
            for f in approved:
                if f not in allowed:
                    continue
                sets.append(f"{f} = :{f}")
                params[f] = cell(row, f)
            if not sets:
                skipped += 1
                continue
            await db.execute(text(
                f"UPDATE public.officials SET {', '.join(sets)} WHERE official_id = :oid"
            ), params)
            merged += 1
            continue

        # action == "insert"
        jid = d.jurisdiction_id
        if not name or not title or not jid:
            errors.append({"row_index": d.row_index, "error": "missing name/title/jurisdiction"})
            continue
        role = cell(row, "role_type") or "elected"
        if role not in ("elected", "staff"):
            role = "elected"
        await db.execute(text(
            "INSERT INTO public.officials "
            "(jurisdiction_id, name, title, email, phone, fax, "
            " mailing_address, physical_address, source, source_date, role_type) "
            "VALUES (:jid, :name, :title, :email, :phone, :fax, "
            "        :mail, :phys, :src, :sd, :role)"
        ), {
            "jid": jid, "name": name, "title": title,
            "email": cell(row, "email"),
            "phone": cell(row, "phone"),
            "fax": cell(row, "fax"),
            "mail": cell(row, "mailing_address"),
            "phys": cell(row, "physical_address"),
            "src": cell(row, "source") or payload.source,
            "sd": today,
            "role": role,
        })
        inserted += 1

    await db.commit()
    return {"inserted": inserted, "merged": merged, "skipped": skipped, "errors": errors}
