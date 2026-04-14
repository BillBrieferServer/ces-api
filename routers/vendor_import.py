from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from pydantic import BaseModel
from typing import Optional
from io import BytesIO
import csv
import re as _re

from openpyxl import load_workbook
from database import get_db

from .import_match import (
    score_company_candidate, score_person_candidate, band, diff_fields,
)

router = APIRouter(prefix="/vendors/import", tags=["vendor-import"])

MAX_PREVIEW_ROWS = 500
MAX_FILE_BYTES = 5_000_000

FIELD_CHOICES = [
    "vendor_name", "contact_name", "contact_title",
    "phone", "cell_phone", "email",
    "website", "address",
    "source", "bluebook_status", "ces_contract_category",
]

GUESS_PATTERNS = {
    "vendor_name": ["vendor", "company", "business", "organization", "org name", "firm"],
    "contact_name": ["contact", "full name", "person", "name"],
    "contact_title": ["title", "position", "role", "job title"],
    "phone": ["work phone", "office phone", "phone", "telephone", "tel"],
    "cell_phone": ["cell", "mobile"],
    "email": ["email", "e-mail", "mail"],
    "website": ["website", "web", "url", "site"],
    "address": ["address", "street", "location"],
    "source": ["source"],
    "bluebook_status": ["bluebook"],
    "ces_contract_category": ["category", "contract"],
}

VENDOR_DIFF_FIELDS = ["vendor_name", "website", "address", "phone", "email",
                      "bluebook_status", "ces_contract_category"]
CONTACT_DIFF_FIELDS = ["contact_name", "contact_title", "phone", "cell_phone", "email"]


def _parse_upload(filename: str, content: bytes):
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None) or []
        headers = [str(h).strip() if h is not None else "" for h in first]
        data = []
        for row in rows_iter:
            padded = list(row) + [None] * (len(headers) - len(row))
            data.append([("" if v is None else str(v)).strip() for v in padded[:len(headers)]])
            if len(data) >= MAX_PREVIEW_ROWS:
                break
        return headers, data
    if name.endswith(".csv") or name.endswith(".txt"):
        text_content = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(text_content.splitlines())
        headers = [h.strip() for h in next(reader, [])]
        data = []
        for row in reader:
            padded = list(row) + [""] * (len(headers) - len(row))
            data.append([c.strip() for c in padded[:len(headers)]])
            if len(data) >= MAX_PREVIEW_ROWS:
                break
        return headers, data
    from .import_vision import extract_via_vision, _detect_media_type
    if _detect_media_type(filename):
        return extract_via_vision(filename, content, "vendor")
    raise HTTPException(400, "Unsupported file type. Use .csv, .xlsx, .pdf, .jpg, or .png")


def _guess_mapping(headers):
    mapping = {}
    used = set()
    for field, pats in GUESS_PATTERNS.items():
        best = None
        for h in headers:
            if h in used or not h:
                continue
            hl = h.lower().strip()
            for p in pats:
                if p == hl:
                    best = h
                    break
            if best:
                break
        if not best:
            for h in headers:
                if h in used or not h:
                    continue
                hl = h.lower().strip()
                if any(p in hl for p in pats):
                    best = h
                    break
        if best:
            mapping[field] = best
            used.add(best)
    return mapping


@router.post("/preview")
async def import_preview(file: UploadFile = File(...)):
    content = await file.read()
    if len(content) > MAX_FILE_BYTES:
        raise HTTPException(413, "File too large (max 5MB)")
    headers, rows = _parse_upload(file.filename, content)
    return {
        "columns": headers,
        "rows": rows,
        "suggested_mapping": _guess_mapping(headers),
        "field_choices": FIELD_CHOICES,
        "truncated": len(rows) >= MAX_PREVIEW_ROWS,
    }


_NAME_SUFFIX_PATTERN = _re.compile(
    r"\b(inc|llc|l\.l\.c|ltd|corp|corporation|company|co|bank|group|services|service|associates|partners|plc)\b\.?",
    _re.IGNORECASE,
)


def _normalize_name(s: str) -> str:
    s = (s or "").lower().strip()
    s = _re.sub(r"^the\s+", "", s)
    s = _NAME_SUFFIX_PATTERN.sub("", s)
    s = _re.sub(r"[^a-z0-9]+", "", s)
    return s


class DiffRequest(BaseModel):
    columns: list[str]
    rows: list[list[str]]
    mapping: dict[str, str]


SIM_THRESHOLD = 0.45


async def _fetch_vendor_candidates(db, vname, incoming):
    """Pull candidate vendors by name similarity + shared email/phone domain."""
    cands = {}

    # Name similarity via pg_trgm
    r = await db.execute(text(
        "SELECT vendor_id, vendor_name, website, address, phone, email, "
        "       bluebook_status, ces_contract_category "
        "FROM ces.vendors "
        "WHERE similarity(lower(vendor_name), :q) > :thr "
        "ORDER BY similarity(lower(vendor_name), :q) DESC LIMIT 10"
    ), {"q": (vname or "").lower(), "thr": SIM_THRESHOLD})
    for v in r.mappings().all():
        cands[v["vendor_id"]] = dict(v)

    # Normalized-name exact match (strips punctuation, LLC, etc.)
    norm = _normalize_name(vname or "")
    if norm:
        r = await db.execute(text(
            "SELECT vendor_id, vendor_name, website, address, phone, email, "
            "       bluebook_status, ces_contract_category FROM ces.vendors"
        ))
        for v in r.mappings().all():
            if _normalize_name(v["vendor_name"]) == norm and v["vendor_id"] not in cands:
                cands[v["vendor_id"]] = dict(v)

    # Also pull by matching email / phone (strong signals)
    inc_email = (incoming.get("email") or "").strip().lower()
    inc_phone = _re.sub(r"\D", "", incoming.get("phone") or "")
    extra_filters = []
    params = {}
    if inc_email:
        extra_filters.append("lower(email) = :em")
        params["em"] = inc_email
    if inc_phone:
        extra_filters.append("regexp_replace(phone, '\\D', '', 'g') = :ph")
        params["ph"] = inc_phone
    if extra_filters:
        q = ("SELECT vendor_id, vendor_name, website, address, phone, email, "
             "bluebook_status, ces_contract_category FROM ces.vendors WHERE "
             + " OR ".join(extra_filters) + " LIMIT 10")
        r = await db.execute(text(q), params)
        for v in r.mappings().all():
            if v["vendor_id"] not in cands:
                cands[v["vendor_id"]] = dict(v)

    return list(cands.values())


async def _fetch_contacts(db, vendor_id):
    r = await db.execute(text(
        "SELECT contact_id, vendor_id, contact_name, contact_title, phone, cell_phone, email, is_primary "
        "FROM ces.vendor_contacts WHERE vendor_id = :vid"
    ), {"vid": vendor_id})
    return [dict(x) for x in r.mappings().all()]


@router.post("/diff")
async def vendor_diff(payload: DiffRequest, db: AsyncSession = Depends(get_db)):
    mapping = payload.mapping or {}
    cols = payload.columns
    col_index = {h: i for i, h in enumerate(cols)}
    if "vendor_name" not in mapping or mapping["vendor_name"] not in col_index:
        raise HTTPException(400, "vendor_name mapping is required")

    def cell(row, field):
        c = mapping.get(field)
        if not c or c not in col_index:
            return None
        i = col_index[c]
        if i >= len(row):
            return None
        v = (row[i] or "").strip()
        return v or None

    results = []

    for idx, row in enumerate(payload.rows):
        vname = cell(row, "vendor_name")
        incoming = {f: cell(row, f) for f in FIELD_CHOICES}

        if not vname:
            results.append({"row_index": idx, "status": "ERROR",
                            "error": "missing vendor_name", "incoming": incoming})
            continue

        cands_raw = await _fetch_vendor_candidates(db, vname, incoming)
        scored = []
        for v in cands_raw:
            score, parts = score_company_candidate(incoming, v)
            scored.append({"score": score, "parts": parts, "vendor": v})
        scored.sort(key=lambda s: s["score"], reverse=True)

        top = scored[0] if scored else None
        top_score = top["score"] if top else 0
        status_band = band(top_score, person=False) if top else "NEW"

        # For top vendor: score contacts
        contact_match = None
        contact_candidates = []
        vendor_fill = []
        vendor_overwrite = []
        has_contact_info = any(incoming.get(f) for f in ["contact_name", "phone", "cell_phone", "email"])

        if top:
            tv = top["vendor"]
            vendor_fill, vendor_overwrite = diff_fields(incoming, tv, VENDOR_DIFF_FIELDS)

            if has_contact_info:
                contacts = await _fetch_contacts(db, tv["vendor_id"])
                inc_contact = {
                    "name": incoming.get("contact_name"),
                    "title": incoming.get("contact_title"),
                    "phone": incoming.get("phone"),
                    "cell_phone": incoming.get("cell_phone"),
                    "email": incoming.get("email"),
                }
                scored_contacts = []
                for c in contacts:
                    existing = {
                        "name": c.get("contact_name"),
                        "title": c.get("contact_title"),
                        "phone": c.get("phone"),
                        "cell_phone": c.get("cell_phone"),
                        "email": c.get("email"),
                    }
                    cs, cp = score_person_candidate(inc_contact, existing, "exact")
                    scored_contacts.append({"score": cs, "parts": cp, "contact": c})
                scored_contacts.sort(key=lambda s: s["score"], reverse=True)
                contact_candidates = [
                    {
                        "contact_id": s["contact"]["contact_id"],
                        "contact_name": s["contact"]["contact_name"],
                        "score": s["score"],
                        "existing": s["contact"],
                    }
                    for s in scored_contacts[:5]
                ]
                if scored_contacts and scored_contacts[0]["score"] >= 70:
                    best_c = scored_contacts[0]
                    mapped_inc = {
                        "contact_name": incoming.get("contact_name"),
                        "contact_title": incoming.get("contact_title"),
                        "phone": incoming.get("phone"),
                        "cell_phone": incoming.get("cell_phone"),
                        "email": incoming.get("email"),
                    }
                    cfill, cover = diff_fields(mapped_inc, best_c["contact"], CONTACT_DIFF_FIELDS)
                    contact_match = {
                        "contact_id": best_c["contact"]["contact_id"],
                        "score": best_c["score"],
                        "existing": best_c["contact"],
                        "fill_fields": cfill,
                        "overwrite_fields": cover,
                    }

        vendor_candidates = [
            {
                "vendor_id": s["vendor"]["vendor_id"],
                "vendor_name": s["vendor"]["vendor_name"],
                "score": s["score"],
                "score_parts": s["parts"],
                "existing": s["vendor"],
            }
            for s in scored[:5]
        ]

        result = {
            "row_index": idx,
            "incoming": incoming,
            "candidates": vendor_candidates,
        }

        if status_band == "AUTO" and top:
            result["status"] = "CHANGED" if (vendor_fill or vendor_overwrite) or contact_match is None and has_contact_info else "AUTO_MATCH"
            result["best_match"] = vendor_candidates[0]
            result["vendor_fill_fields"] = vendor_fill
            result["vendor_overwrite_fields"] = vendor_overwrite
            result["contact_match"] = contact_match
            result["contact_candidates"] = contact_candidates
            result["has_contact_info"] = has_contact_info
        elif status_band == "POSSIBLE" and top:
            result["status"] = "POSSIBLE"
            result["best_match"] = vendor_candidates[0]
            result["vendor_fill_fields"] = vendor_fill
            result["vendor_overwrite_fields"] = vendor_overwrite
            result["contact_candidates"] = contact_candidates
            result["has_contact_info"] = has_contact_info
        else:
            result["status"] = "NEW"

        results.append(result)

    summary = {}
    for r in results:
        summary[r["status"]] = summary.get(r["status"], 0) + 1
    return {"rows": results, "summary": summary}


class CommitDecision(BaseModel):
    row_index: int
    action: str  # "skip" | "insert" | "merge"
    existing_vendor_id: Optional[int] = None
    approved_vendor_fields: Optional[list[str]] = None
    contact_action: Optional[str] = None  # "skip" | "insert" | "merge"
    existing_contact_id: Optional[int] = None
    approved_contact_fields: Optional[list[str]] = None


class CommitRequest(BaseModel):
    columns: list[str]
    rows: list[list[str]]
    mapping: dict[str, str]
    decisions: list[CommitDecision]
    source: Optional[str] = None


@router.post("/commit")
async def vendor_commit(payload: CommitRequest, db: AsyncSession = Depends(get_db)):
    mapping = payload.mapping or {}
    cols = payload.columns
    col_index = {h: i for i, h in enumerate(cols)}

    def cell(row, field):
        c = mapping.get(field)
        if not c or c not in col_index:
            return None
        i = col_index[c]
        if i >= len(row):
            return None
        v = (row[i] or "").strip()
        return v or None

    inserted_vendors = 0
    merged_vendors = 0
    inserted_contacts = 0
    merged_contacts = 0
    skipped = 0
    errors = []

    for d in payload.decisions:
        if d.action == "skip":
            skipped += 1
            continue
        if d.row_index < 0 or d.row_index >= len(payload.rows):
            errors.append({"row_index": d.row_index, "error": "row out of range"})
            continue
        row = payload.rows[d.row_index]
        vname = cell(row, "vendor_name")
        if not vname and d.action == "insert":
            errors.append({"row_index": d.row_index, "error": "missing vendor_name"})
            continue

        contact_payload = {
            "cn": cell(row, "contact_name"),
            "ct": cell(row, "contact_title"),
            "ph": cell(row, "phone"),
            "cp": cell(row, "cell_phone"),
            "em": cell(row, "email"),
        }
        has_contact = any(contact_payload.values())

        vendor_id = None

        if d.action == "merge":
            if not d.existing_vendor_id:
                errors.append({"row_index": d.row_index, "error": "merge needs existing_vendor_id"})
                continue
            vendor_id = d.existing_vendor_id

            # Apply approved vendor field updates
            approved = [f for f in (d.approved_vendor_fields or []) if f in VENDOR_DIFF_FIELDS]
            if approved:
                sets = []
                params = {"vid": vendor_id}
                for f in approved:
                    sets.append(f"{f} = :{f}")
                    params[f] = cell(row, f)
                await db.execute(text(
                    f"UPDATE ces.vendors SET {', '.join(sets)} WHERE vendor_id = :vid"
                ), params)
            merged_vendors += 1

        else:  # insert new vendor
            ins = await db.execute(text(
                "INSERT INTO ces.vendors "
                "(vendor_name, website, address, bluebook_status, ces_contract_category, source, "
                " contact_name, contact_title, phone, cell_phone, email) "
                "VALUES (:vn, :web, :addr, :bb, :cat, :src, :cn, :ct, :ph, :cp, :em) "
                "RETURNING vendor_id"
            ), {
                "vn": vname,
                "web": cell(row, "website"),
                "addr": cell(row, "address"),
                "bb": cell(row, "bluebook_status") or "not_listed",
                "cat": cell(row, "ces_contract_category"),
                "src": cell(row, "source") or payload.source,
                **contact_payload,
            })
            vendor_id = ins.mappings().first()["vendor_id"]
            inserted_vendors += 1

        # Handle contact
        contact_action = d.contact_action or ("insert" if has_contact and d.action == "insert" else None)
        if contact_action == "merge":
            if not d.existing_contact_id:
                errors.append({"row_index": d.row_index, "error": "contact merge needs existing_contact_id"})
                continue
            approved = [f for f in (d.approved_contact_fields or []) if f in CONTACT_DIFF_FIELDS]
            if approved:
                sets = []
                params = {"cid": d.existing_contact_id}
                for f in approved:
                    sets.append(f"{f} = :{f}")
                    params[f] = cell(row, f)
                await db.execute(text(
                    f"UPDATE ces.vendor_contacts SET {', '.join(sets)} WHERE contact_id = :cid"
                ), params)
            merged_contacts += 1
        elif contact_action == "insert" and has_contact:
            is_primary = (d.action == "insert")
            await db.execute(text(
                "INSERT INTO ces.vendor_contacts "
                "(vendor_id, contact_name, contact_title, phone, cell_phone, email, is_primary) "
                "VALUES (:vid, :cn, :ct, :ph, :cp, :em, :pr)"
            ), {"vid": vendor_id, "pr": is_primary, **contact_payload})
            inserted_contacts += 1

    await db.commit()
    return {
        "inserted_vendors": inserted_vendors,
        "merged_vendors": merged_vendors,
        "inserted_contacts": inserted_contacts,
        "merged_contacts": merged_contacts,
        "skipped": skipped,
        "errors": errors,
    }
